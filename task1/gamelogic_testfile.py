import requests
from lxml import html
from random import randint
import openpyxl


# scrapes website for words and saves them into excel worksheet
def get_words():
    with requests.get("https://www.ef.edu/english-resources/english-vocabulary/top-1000-words/") as response:
        tree = html.fromstring(response.content)
        element = tree.xpath(
            "/html/body/div[1]/main/div/div[1]/div/div/div[2]/div/div/div/div/div/section/div/section/article/div/div/div/div/p[2]/text()")
        words = []
        for e in element:
            if len(e) > 0:
                words.append(f"{e}".strip())
        try:
            wb = openpyxl.load_workbook('words.xlsx')
            worksheet = wb.active
            index = 0
            for i in range(1, 1001):
                cell = worksheet.cell(row=i, column=1)
                cell.value = words[index]
                index += 1
            wb.save('words.xlsx')

        except FileNotFoundError:
            wb = openpyxl.Workbook()
            worksheet = wb.active
            worksheet.title = "WordSheet"
            wb.save('words.xlsx')
            # implement for loop from A1 to A1000
            index = 0
            for i in range(1, 1001):
                cell = worksheet.cell(row=i, column=1)
                cell.value = words[index]
                index += 1
            wb.save('words.xlsx')


# selects a random word from the excel worksheet
def select_word():
    random_index = randint(1, 1000)
    wb = openpyxl.load_workbook('words.xlsx')
    worksheet = wb.active
    cell = worksheet.cell(row=random_index, column=1)
    return cell.value


# checks if the guessed word is valid
def check_word(current, word, letter):
    word_list = list(word)
    current_list = list(current)

    if letter in word:
        for index, value in enumerate(word_list):
            if value == letter:
                current_list[index] = letter
        return "".join(current_list)
    else:
        return False


# prompts the user to choose a letter
def choose_letter(guesses):
    if len(guesses) > 0:
        letter = input(f"Choose a letter (guesses: {guesses}): ")
    else:
        letter = input("Choose a letter: ")
    while letter in guesses or len(letter) != 1:
        letter = input("You already guessed this letter or input something invalid, try another one: ")
    guesses.append(letter)
    return letter, guesses


# controls game logic
def start_game(word):
    current = "_" * len(word)
    mistakes = 0
    guesses = []

    while '_' in current and mistakes < 5:
        print(current)
        letter, guesses = choose_letter(guesses)
        result = check_word(current, word, letter)
        if not result:
            mistakes += 1
            print("Incorrect!")
        else:
            current = result
            print(f"Correct!")
    if mistakes < 5:
        print(f"You won in {len(guesses)} guesses! The word was {word}")
    else:
        print(f"Better luck next time, the word was {word}")


# runs main functions needed for the game
def main():
    get_words()
    word = select_word()
    start_game(word)


if __name__ == "__main__":
    main()
